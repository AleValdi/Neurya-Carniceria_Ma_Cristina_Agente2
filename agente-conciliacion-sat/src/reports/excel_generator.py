"""
Generador de reportes Excel para conciliación SAT-ERP
"""
from pathlib import Path
from datetime import datetime
from typing import List, Optional, Dict, Any
from decimal import Decimal
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from loguru import logger

from config.settings import settings
from src.erp.models import ResultadoConciliacion
from src.conciliacion.alerts import Alerta, TipoAlerta


class ExcelReportGenerator:
    """Generador de reportes Excel"""

    # Colores para estilos
    COLOR_HEADER = "1F4E79"
    COLOR_EXITO = "C6EFCE"
    COLOR_ALERTA = "FFEB9C"
    COLOR_ERROR = "FFC7CE"
    COLOR_GRIS = "F2F2F2"

    def __init__(self, output_dir: Optional[Path] = None):
        self.output_dir = output_dir or settings.output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generar_reporte(
        self,
        resultados: List[ResultadoConciliacion],
        alertas: Optional[List[Alerta]] = None,
        nombre_archivo: Optional[str] = None,
        facturas_ya_consolidadas: Optional[List] = None
    ) -> Path:
        """
        Generar reporte Excel completo de conciliación

        Args:
            resultados: Lista de resultados de conciliación
            alertas: Lista de alertas adicionales
            nombre_archivo: Nombre del archivo (sin extensión)
            facturas_ya_consolidadas: Lista de tuplas (Factura, NumRec) omitidas por ya existir

        Returns:
            Ruta al archivo generado
        """
        # Generar nombre de archivo
        if nombre_archivo:
            filename = f"{nombre_archivo}.xlsx"
        else:
            fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{settings.nombre_reporte}_{fecha}.xlsx"

        filepath = self.output_dir / filename

        # Crear workbook
        wb = Workbook()

        # Hoja 1: Resumen ejecutivo
        self._crear_hoja_resumen(wb, resultados, facturas_ya_consolidadas)

        # Hoja 2: Conciliaciones exitosas
        self._crear_hoja_exitosas(wb, resultados)

        # Hoja 3: Conciliaciones con diferencias
        self._crear_hoja_diferencias(wb, resultados)

        # Hoja 4: Facturas sin remisión
        self._crear_hoja_sin_remision(wb, resultados)

        # Hoja 5: Todas las alertas
        self._crear_hoja_alertas(wb, resultados, alertas)

        # Hoja 6: Detalle completo
        self._crear_hoja_detalle(wb, resultados)

        # Hoja 7: Ya consolidadas (si hay)
        if facturas_ya_consolidadas:
            self._crear_hoja_ya_consolidadas(wb, facturas_ya_consolidadas)

        # Eliminar hoja por defecto si existe
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Guardar archivo
        wb.save(filepath)
        logger.info(f"Reporte generado: {filepath}")

        return filepath

    def _crear_hoja_resumen(
        self,
        wb: Workbook,
        resultados: List[ResultadoConciliacion],
        facturas_ya_consolidadas: Optional[List] = None
    ):
        """Crear hoja de resumen ejecutivo"""
        ws = wb.create_sheet("Resumen Ejecutivo", 0)

        # Título
        ws['A1'] = "REPORTE DE CONCILIACIÓN SAT - ERP"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
        ws.merge_cells('A1:E1')

        # Fecha de generación
        ws['A2'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)

        # Calcular estadísticas
        total = len(resultados)
        exitosos = sum(1 for r in resultados if r.conciliacion_exitosa)
        con_diferencias = sum(1 for r in resultados if r.remision and not r.conciliacion_exitosa)
        sin_remision = sum(1 for r in resultados if not r.remision)
        ya_consolidadas = len(facturas_ya_consolidadas) if facturas_ya_consolidadas else 0

        # Tabla de resumen
        row = 4
        headers = ['Métrica', 'Cantidad', 'Porcentaje']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")

        total_xmls = total + ya_consolidadas
        metricas = [
            ('Total XMLs en carpeta', total_xmls, '100%'),
            ('Ya consolidadas (omitidas)', ya_consolidadas, f'{(ya_consolidadas/total_xmls*100) if total_xmls else 0:.1f}%'),
            ('Procesadas', total, f'{(total/total_xmls*100) if total_xmls else 0:.1f}%'),
            ('Conciliaciones exitosas', exitosos, f'{(exitosos/total*100) if total else 0:.1f}%'),
            ('Con diferencias', con_diferencias, f'{(con_diferencias/total*100) if total else 0:.1f}%'),
            ('Sin remisión asociada', sin_remision, f'{(sin_remision/total*100) if total else 0:.1f}%'),
        ]

        for i, (metrica, cantidad, porcentaje) in enumerate(metricas, 1):
            ws.cell(row=row + i, column=1, value=metrica)
            ws.cell(row=row + i, column=2, value=cantidad)
            ws.cell(row=row + i, column=3, value=porcentaje)

            # Color según tipo
            if 'exitosas' in metrica.lower():
                fill = PatternFill(start_color=self.COLOR_EXITO, fill_type="solid")
            elif 'diferencias' in metrica.lower():
                fill = PatternFill(start_color=self.COLOR_ALERTA, fill_type="solid")
            elif 'sin remisión' in metrica.lower():
                fill = PatternFill(start_color=self.COLOR_ERROR, fill_type="solid")
            else:
                fill = PatternFill(start_color=self.COLOR_GRIS, fill_type="solid")

            for col in range(1, 4):
                ws.cell(row=row + i, column=col).fill = fill

        # Monto total procesado
        total_monto = sum(r.total_factura for r in resultados)
        row += len(metricas) + 2
        ws.cell(row=row, column=1, value="Monto total procesado:")
        ws.cell(row=row, column=2, value=f"${float(total_monto):,.2f}")
        ws.cell(row=row, column=1).font = Font(bold=True)

        # Diferencia promedio
        diferencias = [r.diferencia_monto for r in resultados if r.diferencia_monto]
        if diferencias:
            promedio = sum(diferencias) / len(diferencias)
            ws.cell(row=row + 1, column=1, value="Diferencia promedio:")
            ws.cell(row=row + 1, column=2, value=f"${float(promedio):,.2f}")

        # Ajustar anchos de columna
        self._ajustar_anchos(ws)

    def _crear_hoja_exitosas(
        self,
        wb: Workbook,
        resultados: List[ResultadoConciliacion]
    ):
        """Crear hoja de conciliaciones exitosas"""
        ws = wb.create_sheet("Conciliaciones Exitosas")

        exitosos = [r for r in resultados if r.conciliacion_exitosa]

        if not exitosos:
            ws['A1'] = "No hay conciliaciones exitosas en este lote"
            return

        # Crear DataFrame
        data = []
        for r in exitosos:
            data.append({
                'NumRec ERP': r.numero_factura_erp or '',
                'UUID Factura': r.uuid_factura[:8] + '...',
                'Identificador': r.identificador_factura,
                'RFC Emisor': r.rfc_emisor,
                'Proveedor': r.nombre_emisor[:30],
                'Fecha Factura': r.fecha_factura.strftime('%Y-%m-%d') if r.fecha_factura else '',
                'Total Factura': float(r.total_factura),
                'No. Remisión': r.numero_remision or '',
                'Total Remisión': float(r.total_remision) if r.total_remision else 0,
                'Score': f"{r.score_matching:.2f}",
                'Método': r.metodo_matching.upper() if hasattr(r, 'metodo_matching') else 'ALGORITMO',
            })

        df = pd.DataFrame(data)
        self._escribir_dataframe(ws, df)
        self._aplicar_estilo_tabla(ws, len(df) + 1, len(df.columns))

    def _crear_hoja_diferencias(
        self,
        wb: Workbook,
        resultados: List[ResultadoConciliacion]
    ):
        """Crear hoja de conciliaciones con diferencias"""
        ws = wb.create_sheet("Con Diferencias")

        con_diferencias = [r for r in resultados if r.remision and not r.conciliacion_exitosa]

        if not con_diferencias:
            ws['A1'] = "No hay conciliaciones con diferencias en este lote"
            return

        data = []
        for r in con_diferencias:
            data.append({
                'UUID Factura': r.uuid_factura[:8] + '...',
                'Identificador': r.identificador_factura,
                'Proveedor': r.nombre_emisor[:30],
                'Total Factura': float(r.total_factura),
                'No. Remisión': r.numero_remision or '',
                'Total Remisión': float(r.total_remision) if r.total_remision else 0,
                'Diferencia $': float(r.diferencia_monto) if r.diferencia_monto else 0,
                'Diferencia %': f"{r.diferencia_porcentaje:.2f}%" if r.diferencia_porcentaje else '',
                'Alertas': '; '.join(r.alertas[:2]) if r.alertas else '',
            })

        df = pd.DataFrame(data)
        self._escribir_dataframe(ws, df)
        self._aplicar_estilo_tabla(ws, len(df) + 1, len(df.columns), color_filas=self.COLOR_ALERTA)

    def _crear_hoja_sin_remision(
        self,
        wb: Workbook,
        resultados: List[ResultadoConciliacion]
    ):
        """Crear hoja de facturas sin remisión"""
        ws = wb.create_sheet("Sin Remisión")

        sin_remision = [r for r in resultados if not r.remision]

        if not sin_remision:
            ws['A1'] = "Todas las facturas tienen remisión asociada"
            ws['A1'].fill = PatternFill(start_color=self.COLOR_EXITO, fill_type="solid")
            return

        data = []
        for r in sin_remision:
            data.append({
                'UUID Factura': r.uuid_factura,
                'Identificador': r.identificador_factura,
                'RFC Emisor': r.rfc_emisor,
                'Proveedor': r.nombre_emisor,
                'Fecha Factura': r.fecha_factura.strftime('%Y-%m-%d') if r.fecha_factura else '',
                'Total': float(r.total_factura),
            })

        df = pd.DataFrame(data)
        self._escribir_dataframe(ws, df)
        self._aplicar_estilo_tabla(ws, len(df) + 1, len(df.columns), color_filas=self.COLOR_ERROR)

    def _crear_hoja_alertas(
        self,
        wb: Workbook,
        resultados: List[ResultadoConciliacion],
        alertas_adicionales: Optional[List[Alerta]] = None
    ):
        """Crear hoja con todas las alertas"""
        ws = wb.create_sheet("Alertas")

        # Recopilar todas las alertas
        todas_alertas = []
        for r in resultados:
            for alerta_texto in r.alertas:
                todas_alertas.append({
                    'UUID Factura': r.uuid_factura[:8] + '...' if r.uuid_factura else '',
                    'No. Remisión': r.numero_remision or '',
                    'Tipo': alerta_texto.split(':')[0] if ':' in alerta_texto else 'OTRO',
                    'Descripción': alerta_texto,
                })

        # Agregar alertas adicionales si existen
        if alertas_adicionales:
            for alerta in alertas_adicionales:
                todas_alertas.append({
                    'UUID Factura': alerta.uuid_factura[:8] + '...' if alerta.uuid_factura else '',
                    'No. Remisión': alerta.numero_remision or '',
                    'Tipo': alerta.tipo.value,
                    'Descripción': alerta.mensaje,
                })

        if not todas_alertas:
            ws['A1'] = "No se generaron alertas"
            ws['A1'].fill = PatternFill(start_color=self.COLOR_EXITO, fill_type="solid")
            return

        df = pd.DataFrame(todas_alertas)
        self._escribir_dataframe(ws, df)

        # Aplicar colores según tipo de alerta
        for row_idx in range(2, len(todas_alertas) + 2):
            tipo = ws.cell(row=row_idx, column=3).value
            if 'CRITICA' in str(tipo):
                color = self.COLOR_ERROR
            elif 'ALTA' in str(tipo):
                color = self.COLOR_ALERTA
            else:
                color = self.COLOR_GRIS

            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color=color, fill_type="solid"
                )

        self._ajustar_anchos(ws)

    def _crear_hoja_detalle(
        self,
        wb: Workbook,
        resultados: List[ResultadoConciliacion]
    ):
        """Crear hoja con detalle completo"""
        ws = wb.create_sheet("Detalle Completo")

        data = []
        for r in resultados:
            data.append({
                'NumRec ERP': r.numero_factura_erp or '',
                'UUID': r.uuid_factura,
                'Identificador': r.identificador_factura,
                'RFC Emisor': r.rfc_emisor,
                'Nombre Emisor': r.nombre_emisor,
                'Fecha Factura': r.fecha_factura.strftime('%Y-%m-%d %H:%M') if r.fecha_factura else '',
                'Total Factura': float(r.total_factura),
                'No. Remisión': r.numero_remision or 'SIN REMISIÓN',
                'Fecha Remisión': r.fecha_remision.strftime('%Y-%m-%d') if r.fecha_remision else '',
                'Total Remisión': float(r.total_remision) if r.total_remision else 0,
                'Diferencia $': float(r.diferencia_monto) if r.diferencia_monto else 0,
                'Diferencia %': r.diferencia_porcentaje if r.diferencia_porcentaje else 0,
                'Score': r.score_matching,
                'Método': r.metodo_matching.upper() if hasattr(r, 'metodo_matching') else 'ALGORITMO',
                'Estatus': r.resumen_estatus,
                'Alertas': len(r.alertas),
                'Fecha Proceso': r.fecha_procesamiento.strftime('%Y-%m-%d %H:%M'),
            })

        df = pd.DataFrame(data)
        self._escribir_dataframe(ws, df)
        self._aplicar_estilo_tabla(ws, len(df) + 1, len(df.columns))

    def _crear_hoja_ya_consolidadas(self, wb: Workbook, facturas_ya_consolidadas: List):
        """Crear hoja de facturas omitidas por ya estar consolidadas en BD"""
        ws = wb.create_sheet("Ya Consolidadas")

        data = []
        for factura, numrec in facturas_ya_consolidadas:
            data.append({
                'NumRec ERP': f"F-{numrec}",
                'UUID': factura.uuid,
                'Serie-Folio': f"{factura.serie or ''}-{factura.folio or ''}".strip('-'),
                'RFC Emisor': factura.rfc_emisor,
                'Proveedor': factura.nombre_emisor[:40] if factura.nombre_emisor else '',
                'Fecha Factura': factura.fecha_emision.strftime('%Y-%m-%d') if factura.fecha_emision else '',
                'Total': float(factura.total),
            })

        df = pd.DataFrame(data)
        self._escribir_dataframe(ws, df)
        self._aplicar_estilo_tabla(ws, len(df) + 1, len(df.columns), color_filas=self.COLOR_GRIS)

    def _escribir_dataframe(self, ws, df: pd.DataFrame):
        """Escribir DataFrame en hoja de Excel"""
        # Headers
        for col_idx, column in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Datos
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if isinstance(value, float):
                    if abs(value) >= 1000:
                        cell.number_format = '#,##0.00'
                    else:
                        cell.number_format = '0.00'

        self._ajustar_anchos(ws)

    def _aplicar_estilo_tabla(
        self,
        ws,
        num_rows: int,
        num_cols: int,
        color_filas: Optional[str] = None
    ):
        """Aplicar estilos de tabla"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in range(1, num_rows + 1):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

                if row > 1 and color_filas:
                    cell.fill = PatternFill(start_color=color_filas, fill_type="solid")
                elif row > 1 and row % 2 == 0:
                    cell.fill = PatternFill(start_color=self.COLOR_GRIS, fill_type="solid")

    def _ajustar_anchos(self, ws):
        """Ajustar ancho de columnas automáticamente"""
        for column_cells in ws.columns:
            length = max(
                len(str(cell.value) if cell.value else "") for cell in column_cells
            )
            length = min(length + 2, 50)  # Máximo 50 caracteres
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

    def generar_csv(
        self,
        resultados: List[ResultadoConciliacion],
        nombre_archivo: Optional[str] = None
    ) -> Path:
        """
        Generar reporte en formato CSV

        Args:
            resultados: Lista de resultados de conciliación
            nombre_archivo: Nombre del archivo (sin extensión)

        Returns:
            Ruta al archivo generado
        """
        if nombre_archivo:
            filename = f"{nombre_archivo}.csv"
        else:
            fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{settings.nombre_reporte}_{fecha}.csv"

        filepath = self.output_dir / filename

        data = [r.to_dict() for r in resultados]
        df = pd.DataFrame(data)
        df.to_csv(filepath, index=False, encoding='utf-8-sig')

        logger.info(f"Reporte CSV generado: {filepath}")
        return filepath
